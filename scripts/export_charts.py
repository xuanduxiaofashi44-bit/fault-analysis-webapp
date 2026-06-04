"""
设备故障分析 - Excel原生图表导出脚本
由 GitHub Actions 调用，接收 JSON 数据，生成带原生图表的 .xlsx 文件

用法: python export_charts.py <input.json> <output.xlsx>
"""
import json
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="F0F4FA", end_color="F0F4FA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


def style_header(ws, row=1, max_col=10):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_width(ws, min_col=1, max_col=10, max_width=50):
    for col in range(min_col, max_col + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            for cell in row:
                if cell is not None:
                    max_len = max(max_len, len(str(cell)))
        ws.column_dimensions[letter].width = min(max_len + 4, max_width)


def create_data_sheet(wb, records):
    """Sheet 1: 筛选明细"""
    ws = wb.create_sheet("筛选明细")
    headers = ["工作表", "日期", "线体", "起始时间", "截止时间",
               "停机时长(min)", "原机器", "设备类型", "问题描述", "责任部门", "责任人"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    for r, rec in enumerate(records, 2):
        ws.cell(row=r, column=1, value=rec.get("sheet", ""))
        ws.cell(row=r, column=2, value=rec.get("date", ""))
        ws.cell(row=r, column=3, value=rec.get("line", ""))
        ws.cell(row=r, column=4, value=rec.get("startTime", ""))
        ws.cell(row=r, column=5, value=rec.get("endTime", ""))
        ws.cell(row=r, column=6, value=rec.get("downtime", 0))
        ws.cell(row=r, column=7, value=rec.get("machine", ""))
        ws.cell(row=r, column=8, value=rec.get("machineType", ""))
        ws.cell(row=r, column=9, value=rec.get("description", ""))
        ws.cell(row=r, column=10, value=rec.get("department", ""))
        ws.cell(row=r, column=11, value=rec.get("owner", ""))
    style_header(ws, row=1, max_col=11)
    auto_width(ws, max_col=11)
    return ws


def create_type_summary_sheet(wb, type_summary, month_label):
    """Sheet 2: 分类汇总 + 停机柏拉图 + MTTR/MTBF"""
    ws = wb.create_sheet("分类汇总")
    headers = ["设备类型", "故障次数", "停机总时长(min)", "占比", "累计占比", "MTTR(min)", "MTBF(h)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)

    for r, row in enumerate(type_summary, 2):
        ws.cell(row=r, column=1, value=row.get("type", ""))
        ws.cell(row=r, column=2, value=row.get("count", 0))
        ws.cell(row=r, column=3, value=row.get("downtime", 0))
        ws.cell(row=r, column=4, value=row.get("share", 0))
        ws.cell(row=r, column=5, value=row.get("cumulativeShare", 0))
        ws.cell(row=r, column=6, value=row.get("mttr", 0))
        ws.cell(row=r, column=7, value=row.get("mtbf", 0))

    style_header(ws, row=1, max_col=7)
    auto_width(ws, max_col=7)

    # 设置百分比格式
    for r in range(2, 2 + len(type_summary)):
        ws.cell(row=r, column=4).number_format = '0.0%'
        ws.cell(row=r, column=5).number_format = '0.0%'

    n = len(type_summary) + 1  # last data row
    if n < 2:
        return ws

    # ---- 停机柏拉图: 柱状图(停机时长) + 折线图(累计占比) 双Y轴 ----
    chart_pareto = BarChart()
    chart_pareto.type = "col"
    chart_pareto.style = 10
    chart_pareto.title = f"{month_label} 停机柏拉图"
    chart_pareto.y_axis.title = "停机时长 (min)"
    chart_pareto.x_axis.title = "设备类型"
    chart_pareto.width = 22
    chart_pareto.height = 14

    cats = Reference(ws, min_col=1, min_row=2, max_row=n)
    data_downtime = Reference(ws, min_col=3, min_row=1, max_row=n)
    chart_pareto.add_data(data_downtime, titles_from_data=True)
    chart_pareto.set_categories(cats)

    # 停机时长数据标签
    chart_pareto.series[0].dLbls = DataLabelList()
    chart_pareto.series[0].dLbls.showVal = True

    # 折线图: 累计占比, 次坐标轴
    line_cum = LineChart()
    line_cum.y_axis.title = "累计占比"
    line_cum.y_axis.numFmt = '0%'
    cum_data = Reference(ws, min_col=5, min_row=1, max_row=n)
    line_cum.add_data(cum_data, titles_from_data=True)
    line_cum.series[0].dLbls = DataLabelList()
    line_cum.series[0].dLbls.showVal = True
    line_cum.series[0].dLbls.numFmt = '0%'

    chart_pareto.y_axis.crosses = "min"
    chart_pareto += line_cum

    ws.add_chart(chart_pareto, "I2")

    # ---- MTTR/MTBF 折线图 ----
    chart_mttr = LineChart()
    chart_mttr.title = f"{month_label} MTTR / MTBF"
    chart_mttr.y_axis.title = "时间"
    chart_mttr.width = 22
    chart_mttr.height = 14
    chart_mttr.style = 10

    mttr_data = Reference(ws, min_col=6, min_row=1, max_row=n)
    mtbf_data = Reference(ws, min_col=7, min_row=1, max_row=n)
    chart_mttr.add_data(mttr_data, titles_from_data=True)
    chart_mttr.add_data(mtbf_data, titles_from_data=True)
    chart_mttr.set_categories(cats)

    chart_mttr.series[0].dLbls = DataLabelList()
    chart_mttr.series[0].dLbls.showVal = True
    chart_mttr.series[1].dLbls = DataLabelList()
    chart_mttr.series[1].dLbls.showVal = True

    ws.add_chart(chart_mttr, "I19")

    return ws


def create_month_trend_sheet(wb, month_summary):
    """Sheet 3: 月度趋势 + 故障推移图"""
    ws = wb.create_sheet("月度趋势")
    headers = ["月份", "天数", "故障次数", "停机总时长(min)",
               "日均故障时长(min/天)", "故障率(%)", "MTTR(min)", "MTBF(h)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)

    for r, row in enumerate(month_summary, 2):
        ws.cell(row=r, column=1, value=row.get("month", ""))
        ws.cell(row=r, column=2, value=row.get("days", 0))
        ws.cell(row=r, column=3, value=row.get("count", 0))
        ws.cell(row=r, column=4, value=row.get("downtime", 0))
        ws.cell(row=r, column=5, value=row.get("dailyDowntime", 0))
        ws.cell(row=r, column=6, value=row.get("faultRate", 0))
        ws.cell(row=r, column=7, value=row.get("mttr", 0))
        ws.cell(row=r, column=8, value=row.get("mtbf", 0))

    style_header(ws, row=1, max_col=8)
    auto_width(ws, max_col=8)

    n = len(month_summary) + 1
    if n < 2:
        return ws

    # ---- 故障推移图: 故障率柱状图 + 停机总时长折线图 双Y轴 ----
    chart_trend = BarChart()
    chart_trend.type = "col"
    chart_trend.style = 10
    chart_trend.title = "月度故障推移"
    chart_trend.y_axis.title = "停机总时长 (min)"
    chart_trend.x_axis.title = "月份"
    chart_trend.width = 22
    chart_trend.height = 14

    cats = Reference(ws, min_col=1, min_row=2, max_row=n)
    downtime_data = Reference(ws, min_col=4, min_row=1, max_row=n)
    chart_trend.add_data(downtime_data, titles_from_data=True)
    chart_trend.set_categories(cats)
    chart_trend.series[0].dLbls = DataLabelList()
    chart_trend.series[0].dLbls.showVal = True

    # 故障率折线(次坐标)
    line_rate = LineChart()
    line_rate.y_axis.title = "故障率 (%)"
    rate_data = Reference(ws, min_col=6, min_row=1, max_row=n)
    line_rate.add_data(rate_data, titles_from_data=True)
    line_rate.series[0].dLbls = DataLabelList()
    line_rate.series[0].dLbls.showVal = True

    chart_trend.y_axis.crosses = "min"
    chart_trend += line_rate

    ws.add_chart(chart_trend, "J2")

    return ws


def create_daily_trend_sheet(wb, daily_summary, month_label):
    """Sheet 4: 每日趋势 (仅当选择了具体月份)"""
    ws = wb.create_sheet(f"{month_label} 每日趋势")
    headers = ["日期", "故障次数", "停机总时长(min)", "故障率(%)", "MTTR(min)", "MTBF(h)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)

    for r, row in enumerate(daily_summary, 2):
        ws.cell(row=r, column=1, value=row.get("day", ""))
        ws.cell(row=r, column=2, value=row.get("count", 0))
        ws.cell(row=r, column=3, value=row.get("downtime", 0))
        ws.cell(row=r, column=4, value=row.get("faultRate", 0))
        ws.cell(row=r, column=5, value=row.get("mttr", 0))
        ws.cell(row=r, column=6, value=row.get("mtbf", 0))

    style_header(ws, row=1, max_col=6)
    auto_width(ws, max_col=6)

    n = len(daily_summary) + 1
    if n < 2:
        return ws

    # 每日故障推移
    chart_daily = BarChart()
    chart_daily.type = "col"
    chart_daily.style = 10
    chart_daily.title = f"{month_label} 每日故障推移"
    chart_daily.y_axis.title = "停机总时长 (min)"
    chart_daily.width = 22
    chart_daily.height = 14

    cats = Reference(ws, min_col=1, min_row=2, max_row=n)
    downtime_data = Reference(ws, min_col=3, min_row=1, max_row=n)
    chart_daily.add_data(downtime_data, titles_from_data=True)
    chart_daily.set_categories(cats)
    chart_daily.series[0].dLbls = DataLabelList()
    chart_daily.series[0].dLbls.showVal = True

    # 故障率折线(次坐标)
    line_daily = LineChart()
    line_daily.y_axis.title = "故障率 (%)"
    rate_data = Reference(ws, min_col=4, min_row=1, max_row=n)
    line_daily.add_data(rate_data, titles_from_data=True)
    line_daily.series[0].dLbls = DataLabelList()
    line_daily.series[0].dLbls.showVal = True

    chart_daily.y_axis.crosses = "min"
    chart_daily += line_daily

    ws.add_chart(chart_daily, "H2")

    return ws


def main():
    if len(sys.argv) != 3:
        print("用法: python export_charts.py <input.json> <output.xlsx>")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2]

    with open(input_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    # 删除默认空sheet
    wb.remove(wb.active)

    month_label = data.get("month", "合计")
    records = data.get("records", [])
    type_summary = data.get("typeSummary", [])
    month_summary = data.get("monthSummary", [])
    daily_summary = data.get("dailySummary", [])

    # Sheet 1: 筛选明细
    if records:
        create_data_sheet(wb, records)

    # Sheet 2: 分类汇总 + 图表
    if type_summary:
        create_type_summary_sheet(wb, type_summary, month_label)

    # Sheet 3: 月度趋势 + 图表
    if month_summary:
        create_month_trend_sheet(wb, month_summary)

    # Sheet 4: 每日趋势 (仅具体月份)
    if daily_summary and month_label != "合计":
        create_daily_trend_sheet(wb, daily_summary, month_label)

    wb.save(output_file)
    print(f"导出完成: {output_file}")


if __name__ == "__main__":
    main()
